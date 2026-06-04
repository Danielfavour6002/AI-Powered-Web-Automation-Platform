"""Enterprise Excel report generation with embedded screenshots and color-coded rows."""

import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from core.models import Run, Result
from core.exceptions import ReportError

logger = logging.getLogger(__name__)

# Colour palette constants
COL_HEADER_BG   = "1E3A5F"   # Deep navy
COL_HEADER_FG   = "FFFFFF"
COL_PASS_BG     = "C6EFCE"
COL_PASS_FG     = "276221"
COL_FAIL_BG     = "FFC7CE"
COL_FAIL_FG     = "9C0006"
COL_SKIP_BG     = "FFEB9C"
COL_SKIP_FG     = "9C6500"
COL_TITLE_BG    = "0070F3"
COL_ACCENT      = "2563EB"
ROW_HEIGHT_SCREENSHOT = 80    # points


def _thin_border() -> Border:
    """Return a uniform thin border object."""
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(cell, text: str) -> None:
    """Apply column-header styling to a cell."""
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color=COL_HEADER_FG, size=10)
    cell.fill = PatternFill(start_color=COL_HEADER_BG, end_color=COL_HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _status_fill(status: str) -> Optional[PatternFill]:
    """Return a PatternFill corresponding to the result status."""
    if status == "passed":
        return PatternFill(start_color=COL_PASS_BG, end_color=COL_PASS_BG, fill_type="solid")
    elif status == "failed":
        return PatternFill(start_color=COL_FAIL_BG, end_color=COL_FAIL_BG, fill_type="solid")
    elif status == "skipped":
        return PatternFill(start_color=COL_SKIP_BG, end_color=COL_SKIP_BG, fill_type="solid")
    return None


def _apply_row_style(ws, row_idx: int, n_cols: int, status: str) -> None:
    """Apply alternating row highlighting based on step status."""
    fill = _status_fill(status)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill:
            cell.fill = fill
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def generate_excel_report(
    run: Run,
    results: List[Result],
    output_path: Path,
    client_code: Optional[str] = None,
) -> Path:
    """
    Generate a professional multi-tab Excel report.

    Tabs:
      1. Executive Summary  – pass counts, duration, consultant info
      2. Steps Detail       – per-step rows with inline embedded screenshots
      3. Metadata           – run & DB references

    Args:
        run: The Run entity whose results are being reported.
        results: Ordered list of Result entities for the run.
        output_path: Destination .xlsx path (parent dirs must exist).
        client_code: Short client identifier used in filename generation.

    Returns:
        Path: The path to the saved workbook.
    """
    try:
        wb = Workbook()

        # ── Sheet 1 : Executive Summary ──────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.sheet_view.showGridLines = False

        # Title banner
        ws_sum.merge_cells("A1:F1")
        tc = ws_sum["A1"]
        tc.value = "QA AUTOMATION — TEST RUN REPORT"
        tc.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        tc.fill = PatternFill(start_color=COL_TITLE_BG, end_color=COL_TITLE_BG, fill_type="solid")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 36

        # Section header
        ws_sum.merge_cells("A2:F2")
        sc = ws_sum["A2"]
        sc.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M UTC')}"
        sc.font = Font(name="Calibri", size=10, italic=True, color="555555")
        sc.alignment = Alignment(horizontal="center")

        # Key-value pairs
        passed_c = run.passed_count
        failed_c = run.failed_count
        skipped_c = max(0, run.step_count - passed_c - failed_c)
        pass_rate = f"{(passed_c / run.step_count * 100):.1f}%" if run.step_count else "N/A"

        info_rows = [
            ("Test Name",    run.test_name),
            ("Run ID",       run.id),
            ("Status",       run.status.value.upper()),
            ("Client Code",  client_code or run.pod or "—"),
            ("Consultant",   run.consultant or "—"),
            ("Started",      run.started_at or run.created_at),
            ("Duration",     f"{run.duration_seconds:.1f}s" if run.duration_seconds else "—"),
            ("Total Steps",  str(run.step_count)),
            ("Passed",       str(passed_c)),
            ("Failed",       str(failed_c)),
            ("Skipped",      str(skipped_c)),
            ("Pass Rate",    pass_rate),
        ]

        r = 4
        for label, value in info_rows:
            lc = ws_sum.cell(row=r, column=1, value=label)
            lc.font = Font(name="Calibri", bold=True, size=10, color=COL_HEADER_BG)
            lc.alignment = Alignment(vertical="center")

            vc = ws_sum.cell(row=r, column=2, value=value)
            vc.font = Font(name="Calibri", size=10)
            vc.alignment = Alignment(vertical="center")

            if label == "Status":
                fill = _status_fill(run.status.value)
                if fill:
                    vc.fill = fill
                if value == "PASSED":
                    vc.font = Font(name="Calibri", bold=True, color=COL_PASS_FG)
                elif value == "FAILED":
                    vc.font = Font(name="Calibri", bold=True, color=COL_FAIL_FG)

            ws_sum.row_dimensions[r].height = 18
            r += 1

        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 50

        # ── Sheet 2 : Steps Detail ────────────────────────────────────────────
        ws_steps = wb.create_sheet(title="Steps Detail")
        ws_steps.sheet_view.showGridLines = False

        col_headers = [
            "Step #", "Action", "Description", "Selector",
            "Expected", "Actual / Error", "Status", "Duration (ms)", "Screenshot",
        ]
        for ci, h in enumerate(col_headers, 1):
            _header_style(ws_steps.cell(row=1, column=ci), h)

        ws_steps.freeze_panes = "A2"
        ws_steps.auto_filter.ref = f"A1:{get_column_letter(len(col_headers))}1"
        ws_steps.row_dimensions[1].height = 24

        # Set screenshot column width & height preset
        screenshot_col = 9
        ws_steps.column_dimensions[get_column_letter(screenshot_col)].width = 22

        for ri, res in enumerate(results, 2):
            status = res.status if isinstance(res.status, str) else res.status.value

            # Row data
            ws_steps.cell(row=ri, column=1, value=res.sequence)
            ws_steps.cell(row=ri, column=2, value=res.action.value if hasattr(res.action, "value") else str(res.action))
            ws_steps.cell(row=ri, column=3, value=res.description or "")
            ws_steps.cell(row=ri, column=4, value=res.selector or "")
            ws_steps.cell(row=ri, column=5, value=res.expected_value or "")

            actual = res.actual_value or ""
            if res.error_message:
                actual = res.error_message[:200]
            ws_steps.cell(row=ri, column=6, value=actual)

            # Status badge cell
            sc = ws_steps.cell(row=ri, column=7, value=status.upper())
            sc.font = Font(name="Calibri", bold=True, size=10)
            if status == "passed":
                sc.font = Font(name="Calibri", bold=True, color=COL_PASS_FG)
            elif status == "failed":
                sc.font = Font(name="Calibri", bold=True, color=COL_FAIL_FG)
            elif status == "skipped":
                sc.font = Font(name="Calibri", bold=True, color=COL_SKIP_FG)

            ws_steps.cell(row=ri, column=8, value=res.duration_ms or 0)

            # Embed screenshot inline in cell
            if res.screenshot_path:
                ss_path = Path(res.screenshot_path)
                if ss_path.exists():
                    try:
                        img = XLImage(str(ss_path))
                        # Scale to fit column height
                        img.height = ROW_HEIGHT_SCREENSHOT
                        img.width = int(img.height * 1.6)
                        cell_ref = f"{get_column_letter(screenshot_col)}{ri}"
                        ws_steps.add_image(img, cell_ref)
                        ws_steps.row_dimensions[ri].height = ROW_HEIGHT_SCREENSHOT
                    except Exception as img_err:
                        logger.warning(f"Could not embed screenshot {ss_path}: {img_err}")
                        ws_steps.cell(row=ri, column=screenshot_col, value=ss_path.name)
                else:
                    ws_steps.cell(row=ri, column=screenshot_col, value="(file missing)")
            else:
                ws_steps.row_dimensions[ri].height = 20

            _apply_row_style(ws_steps, ri, len(col_headers), status)

        # Column widths for steps sheet
        step_widths = [8, 14, 40, 40, 24, 36, 12, 14, 22]
        for ci, w in enumerate(step_widths, 1):
            ws_steps.column_dimensions[get_column_letter(ci)].width = w

        # ── Sheet 3 : Metadata ────────────────────────────────────────────────
        ws_meta = wb.create_sheet(title="Metadata")
        ws_meta.sheet_view.showGridLines = False

        meta_pairs = [
            ("run_id",       run.id),
            ("test_id",      run.test_id),
            ("client_id",    getattr(run, "client_id", "") or ""),
            ("run_dir",      run.run_dir or ""),
            ("video_path",   run.video_path or ""),
            ("trace_path",   run.trace_path or ""),
            ("created_at",   run.created_at),
            ("completed_at", run.completed_at or ""),
            ("error",        run.error_message or ""),
        ]
        for mr, (k, v) in enumerate(meta_pairs, 1):
            ws_meta.cell(row=mr, column=1, value=k).font = Font(bold=True, color=COL_HEADER_BG)
            ws_meta.cell(row=mr, column=2, value=str(v))
        ws_meta.column_dimensions["A"].width = 20
        ws_meta.column_dimensions["B"].width = 80

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info(f"Excel report saved: {output_path}")
        return output_path

    except Exception as exc:
        raise ReportError(f"Failed to generate Excel report: {exc}") from exc


def build_report_filename(client_code: str, test_name: str, timestamp: Optional[str] = None) -> str:
    """
    Build a standardised report filename.

    Pattern: <ClientCode>_<SanitisedTestName>_<RunTimestamp>.xlsx

    Args:
        client_code: Short client identifier (e.g. 'ACME').
        test_name: The name of the test being reported.
        timestamp: Optional ISO-like timestamp string; defaults to now.

    Returns:
        str: Filename without directory prefix.
    """
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in test_name)[:40]
    safe_code = "".join(c if c.isalnum() else "_" for c in client_code)[:10]
    return f"{safe_code}_{safe_name}_{ts}.xlsx"
